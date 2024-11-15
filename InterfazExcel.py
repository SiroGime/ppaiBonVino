import openpyxl
from openpyxl import Workbook

class InterfazExcel:
    def exportar_excel(self, vinos):
        # Crear un nuevo libro de trabajo
        wb = Workbook()

        # Seleccionar la hoja de trabajo activa
        ws = wb.active

        # Escribir los encabezados en la hoja
        ws['A1'] = 'Nombre del Vino'
        ws['B1'] = 'Bodega'
        ws['C1'] = 'Región'
        ws['D1'] = 'País'
        ws['E1'] = 'Varietal'
        ws['F1'] = 'Promedio de Puntaje de Reseña'
        ws['G1'] = 'Precio Sugerido' 

        # Agregar los datos de los vinos
        for idx, vino in enumerate(vinos, start=2):
            ws[f'A{idx}'] = vino['nombre_vino']
            ws[f'B{idx}'] = vino['bodega']
            ws[f'C{idx}'] = vino['region']
            ws[f'D{idx}'] = vino['pais']
            ws[f'E{idx}'] = vino['varietal_descripcion']
            ws[f'F{idx}'] =  vino['Promedio de la Calificacion de reseña']
            ws[f'G{idx}'] = vino['precio_sugerido']

        # Guardar el archivo
        wb.save("ranking_vinos.xlsx")

        print("Archivo Excel creado exitosamente.")

        return True
